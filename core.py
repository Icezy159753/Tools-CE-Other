"""
core.py — Other Recode Tool: Business Logic (no GUI, no CLI)
All heavy lifting lives here so the GUI stays thin.
"""

from __future__ import annotations

import json
import logging
import numbers
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from urllib import error as urlerror
from urllib import request as urlrequest

import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants (editable here; GUI can override via function args)
# ---------------------------------------------------------------------------
OTHER_LABEL_KEYWORDS: list[str] = [
    "อื่น", "ระบุ", "other", "specify", "others", "else",
]
OTH_SUFFIX: str = "_oth"
NEW_CODE_COL: str = "New_Code"
NEW_OPEN_TEXT_COL: str = "New_Open_Text"
SBJNUM_COL: str = "Sbjnum"
CODING_COLUMNS: list[str] = [
    "Question",
    "Variable_Label",
    SBJNUM_COL,
    "Other_Label",
    "Other_Code",
    NEW_CODE_COL,
    "Open_Text",
    NEW_OPEN_TEXT_COL,
    "Open_Text_From",
    "Remark",
]
MA_CODES_COL: str = "MA_Answers"
CODEFRAME_COLUMNS: list[str] = [
    "Source_Group",
    "Variable_Label",
    "Code No.",
    "Thai Group1",
    "Thai Group2",
    "English",
    "Count",
    "AI_Model",
]
OPENROUTER_BASE_URL: str = "https://openrouter.ai/api/v1/chat/completions"
OPENROUTER_MODEL_DEFAULT: str = "openrouter/auto"
OPENROUTER_MAX_WORKERS: int = 3

_OTHER_PATTERN = re.compile(
    "|".join(re.escape(k) for k in OTHER_LABEL_KEYWORDS),
    flags=re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Result dataclasses (plain dicts for simplicity)
# ---------------------------------------------------------------------------

class Phase1Result:
    """Return value from phase1_export()."""
    def __init__(
        self,
        coding_df: pd.DataFrame,
        output_path: Path,
        n_questions: int,
        n_rows: int,
        detected: dict[str, list[str]],   # {q_col: [other_codes]}
    ) -> None:
        self.coding_df = coding_df
        self.output_path = output_path
        self.n_questions = n_questions
        self.n_rows = n_rows
        self.detected = detected           # for display in GUI


class Phase2Result:
    """Return value from phase2_apply()."""
    def __init__(
        self,
        log_df: pd.DataFrame,
        skipped_df: pd.DataFrame,
        not_found_df: pd.DataFrame,
        output_rawdata_path: Path,
        n_applied: int,
        n_skipped: int,
        n_not_found: int,
    ) -> None:
        self.log_df = log_df
        self.skipped_df = skipped_df
        self.not_found_df = not_found_df
        self.output_rawdata_path = output_rawdata_path
        self.n_applied = n_applied
        self.n_skipped = n_skipped
        self.n_not_found = n_not_found


class CodeFrameResult:
    """Return value from codeframe generation."""
    def __init__(
        self,
        codeframe_df: pd.DataFrame,
        output_path: Path,
        n_groups: int,
        n_rows: int,
        model: str,
    ) -> None:
        self.codeframe_df = codeframe_df
        self.output_path = output_path
        self.n_groups = n_groups
        self.n_rows = n_rows
        self.model = model


class ColumnMappingReport:
    """Column mapping report between Excel rawdata columns and SPSS variables."""

    def __init__(
        self,
        resolved_var_labels: dict[str, str],
        resolved_value_labels: dict[str, dict],
        direct_matches: list[tuple[str, str]],
        fallback_matches: list[tuple[str, str]],
        mismatched_matches: list[tuple[str, str]],
        unresolved_excel_cols: list[str],
    ) -> None:
        self.resolved_var_labels = resolved_var_labels
        self.resolved_value_labels = resolved_value_labels
        self.direct_matches = direct_matches
        self.fallback_matches = fallback_matches
        self.mismatched_matches = mismatched_matches
        self.unresolved_excel_cols = unresolved_excel_cols


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def read_rawdata(excel_path: Path, as_str: bool = True) -> pd.DataFrame:
    """Read the first sheet of an Excel rawdata file."""
    logger.info(f"Reading rawdata: {excel_path.name}")
    if as_str:
        df = pd.read_excel(excel_path, dtype=str)
    else:
        # Keep original types to avoid writing numeric-looking values as text.
        df = pd.read_excel(excel_path)
    logger.info(f"  → {len(df)} rows, {len(df.columns)} cols")
    return df


def find_oth_pairs(df: pd.DataFrame) -> list[tuple[str, str]]:
    """Return direct pairs [(q_col, q_oth_col), ...] (no cartesian expansion)."""
    pairs: list[tuple[str, str]] = []
    cols = [str(c) for c in df.columns]
    cols_set = set(cols)
    norm_map = {_norm_col_name(c): c for c in cols}

    for col in cols:
        q_col: str | None = None

        # Pattern 1: <question>_oth
        if re.search(r"_oth$", col, flags=re.IGNORECASE):
            base = col[: -len(OTH_SUFFIX)]
            if base in cols_set:
                q_col = base

        # Pattern 2 (direct only): <prefix>_oth1 -> <prefix>_O1 / <prefix>_O01
        prefix, idx_int = _parse_oth_col_name(col)
        if q_col is None and prefix is not None and idx_int is not None:
            candidates = [
                f"{prefix}_O{idx_int}",
                f"{prefix}_O{idx_int:02d}",
                f"{prefix}_o{idx_int}",
                f"{prefix}_o{idx_int:02d}",
                f"{prefix}_{idx_int}",
                f"{prefix}_{idx_int:02d}",
            ]
            for cand in candidates:
                if cand in cols_set:
                    q_col = cand
                    break
                cand_norm = norm_map.get(_norm_col_name(cand))
                if cand_norm:
                    q_col = cand_norm
                    break

        if q_col:
            pairs.append((q_col, col))

    # keep unique order
    seen: set[tuple[str, str]] = set()
    uniq_pairs: list[tuple[str, str]] = []
    for p in pairs:
        if p not in seen:
            uniq_pairs.append(p)
            seen.add(p)

    logger.info(f"  -> {len(uniq_pairs)} pair(s): {[p[0] for p in uniq_pairs]}")
    return uniq_pairs


def read_spss_labels(spss_path: Path) -> tuple[dict[str, str], dict[str, dict]]:
    """Return (variable_labels, value_labels) from a .sav file."""
    logger.info(f"Reading SPSS: {spss_path.name}")
    _, meta = pyreadstat.read_sav(
        str(spss_path),
        apply_value_formats=False,
        formats_as_category=False,
    )
    var_labels = dict(zip(meta.column_names, meta.column_labels, strict=False))
    return var_labels, meta.variable_value_labels


def detect_other_codes(q_col: str, value_labels: dict[str, dict]) -> list[str]:
    """Return all codes in SPSS value labels that match OTHER keywords."""
    vl = _get_value_labels_for_qcol(q_col, value_labels)
    matched: list[str] = []
    for code, label in vl.items():
        if _OTHER_PATTERN.search(str(label)):
            matched.append(_normalize_code_str(code))
    return matched


def _build_labels_ref(vl: dict) -> str:
    if not vl:
        return "(no labels)"
    return ", ".join(
        f"{int(k) if isinstance(k, float) else k}={v}"
        for k, v in sorted(
            vl.items(),
            key=lambda x: float(x[0]) if str(x[0]).replace(".", "").isdigit() else 0,
        )
    )


def _get_label(vl: dict, code_str: str) -> str:
    target = _normalize_code_str(code_str)
    for k, v in vl.items():
        if _normalize_code_str(k) == target:
            return str(v)
    return code_str


def _norm_col_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def _has_oth_marker(name: str) -> bool:
    """Return True when the column name contains 'oth' in any casing/format."""
    return "oth" in _norm_col_name(name)


def _parse_oth_col_name(name: str) -> tuple[str | None, int | None]:
    """Parse oth-like column names into (family_prefix, code_index)."""
    raw = str(name).strip()
    patterns = [
        r"^(.*)_oth(\d+)$",
        r"^(.*)_(\d+)_oth$",
        r"^(.*)oth(\d+)$",
        r"^(.*?)(\d+)oth$",
    ]
    for pattern in patterns:
        match = re.match(pattern, raw, flags=re.IGNORECASE)
        if match:
            prefix = str(match.group(1)).rstrip("_")
            return prefix, int(match.group(2))
    if re.search(r"_oth$", raw, flags=re.IGNORECASE):
        return raw[: -len(OTH_SUFFIX)], None
    if _has_oth_marker(raw):
        return raw, None
    return None, None


def _parse_question_family(name: str) -> tuple[str | None, int | None]:
    """Parse question-like column names into (family_prefix, code_index)."""
    raw = str(name).strip()
    match = re.match(r"^(.*)_(?:o)?(\d+)$", raw, flags=re.IGNORECASE)
    if match and not _has_oth_marker(raw):
        return str(match.group(1)), int(match.group(2))
    return None, None


def _normalize_code_str(value) -> str:
    s = str(value).strip()
    if re.fullmatch(r"[+-]?\d+\.0+", s):
        return str(int(float(s)))
    return s


def _infer_oth_col_for_question(q_col: str, all_cols: list[str]) -> str | None:
    """Infer open-text column for a question (supports f3_O3 -> f3_oth style)."""
    cols_set = set(all_cols)
    norm_map = {_norm_col_name(c): c for c in all_cols}

    # Exact pair style: q -> q_oth
    direct = f"{q_col}{OTH_SUFFIX}"
    if direct in cols_set:
        return direct
    direct_norm = norm_map.get(_norm_col_name(direct))
    if direct_norm:
        return direct_norm
    for col in all_cols:
        if _has_oth_marker(col) and _norm_col_name(col).startswith(_norm_col_name(q_col)):
            return col

    # MA family style: f3_O3 -> f3_oth
    m = re.match(r"^(.*)_o\d+$", q_col, flags=re.IGNORECASE)
    if m:
        prefix = m.group(1)
        family_oth = f"{prefix}{OTH_SUFFIX}"
        if family_oth in cols_set:
            return family_oth
        family_norm = norm_map.get(_norm_col_name(family_oth))
        if family_norm:
            return family_norm
        for col in all_cols:
            if _has_oth_marker(col) and _norm_col_name(col).startswith(_norm_col_name(prefix)):
                return col

    return None


def _infer_q_candidates_for_oth(oth_col: str, all_cols: list[str]) -> list[str]:
    """Infer possible question columns for any oth-style column."""
    cols_set = set(all_cols)
    norm_map = {_norm_col_name(c): c for c in all_cols}
    out: list[str] = []

    def _add(cand: str) -> None:
        if cand in cols_set and cand not in out:
            out.append(cand)
            return
        norm_hit = norm_map.get(_norm_col_name(cand))
        if norm_hit and norm_hit not in out:
            out.append(norm_hit)

    prefix, idx = _parse_oth_col_name(oth_col)
    if prefix is not None and idx is not None:
        _add(f"{prefix}_O{idx}")
        _add(f"{prefix}_O{idx:02d}")
        _add(f"{prefix}_{idx}")
        _add(f"{prefix}_{idx:02d}")
        # Indexed oth must stick to same indexed question only.
        return out
    if prefix is not None and idx is None and re.search(r"_oth$", oth_col, flags=re.IGNORECASE):
        _add(prefix)
        fam_q: list[tuple[int, str]] = []
        for c in all_cols:
            q_prefix, q_idx = _parse_question_family(c)
            if q_prefix == prefix and q_idx is not None:
                fam_q.append((q_idx, c))
        for _, c in sorted(fam_q, key=lambda x: x[0]):
            if c not in out:
                out.append(c)
        return out
    elif re.search(r"_oth$", oth_col, flags=re.IGNORECASE):
        prefix = oth_col[: -len(OTH_SUFFIX)]
        _add(prefix)
    else:
        prefix = oth_col

    fam_q: list[tuple[int, str]] = []
    for c in all_cols:
        q_prefix, q_idx = _parse_question_family(c)
        if q_prefix == prefix and q_idx is not None:
            fam_q.append((q_idx, c))
    for _, c in sorted(fam_q, key=lambda x: x[0]):
        if c not in out:
            out.append(c)

    return out


def _get_value_labels_for_qcol(q_col: str, value_labels: dict[str, dict]) -> dict:
    """Get SPSS value labels for q_col with tolerant name matching."""
    if q_col in value_labels:
        return value_labels[q_col]

    # e.g. s10_1_O1 <-> s10_1_01
    candidates = [q_col]
    m_o = re.match(r"^(.*)_o0*(\d+)$", q_col, flags=re.IGNORECASE)
    if m_o:
        prefix, idx_str = m_o.group(1), m_o.group(2)
        idx_int = int(idx_str)
        candidates.extend([f"{prefix}_{idx_int}", f"{prefix}_{idx_int:02d}"])
    m_n = re.match(r"^(.*)_0*(\d+)$", q_col, flags=re.IGNORECASE)
    if m_n:
        prefix, idx_str = m_n.group(1), m_n.group(2)
        idx_int = int(idx_str)
        candidates.extend([f"{prefix}_O{idx_int}", f"{prefix}_O{idx_int:02d}"])
    for cand in candidates:
        if cand in value_labels:
            return value_labels[cand]

    target_norm = _norm_col_name(q_col)
    for key, vl in value_labels.items():
        if _norm_col_name(key) == target_norm:
            return vl
    return {}


def _resolve_sbjnum_col(df: pd.DataFrame) -> str | None:
    """Find Sbjnum-like column name with loose matching."""
    norm_map = {_norm_col_name(col): str(col) for col in df.columns}
    candidates = ("sbjnum", "subjectnumber", "respondentid", "id")
    for key in candidates:
        if key in norm_map:
            return norm_map[key]
    return None


def _column_tail_token(name: str) -> str:
    """Return the most specific trailing token for fuzzy Excel/SPSS column matching."""
    parts = [p for p in re.split(r"[_\W]+", str(name).strip(), flags=re.UNICODE) if p]
    if not parts:
        return _norm_col_name(name)
    return _norm_col_name(parts[-1])


def _build_spss_column_mapping(
    excel_columns: list[str],
    variable_labels: dict[str, str],
    value_labels: dict[str, dict],
    allow_order_fallback: bool = False,
) -> ColumnMappingReport:
    """Map Excel columns to SPSS variable names using exact/tolerant match, then optional order fallback."""
    spss_cols = list(variable_labels.keys())
    spss_set = set(spss_cols)
    spss_norm_map = {_norm_col_name(col): col for col in spss_cols}
    spss_tail_map: dict[str, list[str]] = {}
    for col in spss_cols:
        spss_tail_map.setdefault(_column_tail_token(col), []).append(col)
    resolved_var_labels: dict[str, str] = {}
    resolved_value_labels: dict[str, dict] = {}
    direct_matches: list[tuple[str, str]] = []
    fallback_matches: list[tuple[str, str]] = []
    mismatched_matches: list[tuple[str, str]] = []
    unresolved_excel_cols: list[str] = []
    used_spss_cols: set[str] = set()

    def _add_mapping(excel_col: str, spss_col: str, is_fallback: bool) -> None:
        resolved_var_labels[excel_col] = variable_labels.get(spss_col, "") or ""
        resolved_value_labels[excel_col] = value_labels.get(spss_col, {})
        used_spss_cols.add(spss_col)
        if is_fallback:
            fallback_matches.append((excel_col, spss_col))
        else:
            direct_matches.append((excel_col, spss_col))
        if str(excel_col) != str(spss_col):
            mismatched_matches.append((excel_col, spss_col))

    for idx, excel_col in enumerate(excel_columns):
        if excel_col in spss_set:
            _add_mapping(excel_col, excel_col, is_fallback=False)
            continue

        norm_hit = spss_norm_map.get(_norm_col_name(excel_col))
        if norm_hit:
            _add_mapping(excel_col, norm_hit, is_fallback=False)
            continue

        # e.g. s10_1_O1 <-> s10_1_01 or inverse
        matched = None
        m_o = re.match(r"^(.*)_o0*(\d+)$", excel_col, flags=re.IGNORECASE)
        if m_o:
            prefix, idx_str = m_o.group(1), m_o.group(2)
            idx_int = int(idx_str)
            for cand in (f"{prefix}_{idx_int}", f"{prefix}_{idx_int:02d}"):
                if cand in spss_set:
                    matched = cand
                    break
        if not matched:
            m_n = re.match(r"^(.*)_0*(\d+)$", excel_col, flags=re.IGNORECASE)
            if m_n:
                prefix, idx_str = m_n.group(1), m_n.group(2)
                idx_int = int(idx_str)
                for cand in (f"{prefix}_O{idx_int}", f"{prefix}_O{idx_int:02d}"):
                    if cand in spss_set:
                        matched = cand
                        break
        if matched:
            _add_mapping(excel_col, matched, is_fallback=False)
            continue

        excel_tail = _column_tail_token(excel_col)
        tail_candidates = [
            cand for cand in spss_tail_map.get(excel_tail, []) if cand not in used_spss_cols
        ]
        if len(tail_candidates) == 1:
            _add_mapping(excel_col, tail_candidates[0], is_fallback=False)
            continue

        if allow_order_fallback and idx < len(spss_cols):
            fallback_target = spss_cols[idx]
            if fallback_target not in used_spss_cols:
                _add_mapping(excel_col, fallback_target, is_fallback=True)
                continue

        unresolved_excel_cols.append(excel_col)

    return ColumnMappingReport(
        resolved_var_labels=resolved_var_labels,
        resolved_value_labels=resolved_value_labels,
        direct_matches=direct_matches,
        fallback_matches=fallback_matches,
        mismatched_matches=mismatched_matches,
        unresolved_excel_cols=unresolved_excel_cols,
    )


def inspect_phase1_column_mapping(
    rawdata_path: Path,
    spss_path: Path,
    allow_order_fallback: bool = False,
) -> ColumnMappingReport:
    """Inspect Excel vs SPSS variable mapping before running phase 1."""
    df = read_rawdata(rawdata_path, as_str=True)
    variable_labels, value_labels = read_spss_labels(spss_path)
    sbjnum_col = _resolve_sbjnum_col(df)
    candidate_cols = [
        str(col)
        for col in df.columns
        if str(col) != (sbjnum_col or "") and not _has_oth_marker(str(col))
    ]
    return _build_spss_column_mapping(
        candidate_cols,
        variable_labels,
        value_labels,
        allow_order_fallback=allow_order_fallback,
    )


def _is_meaningful_ma_value(value) -> bool:
    """Return True when the MA cell represents a selected answer."""
    if _is_empty_value(value):
        return False
    normalized = _normalize_code_str(value)
    return normalized not in {"0", "0.0"}


def _get_ma_family_columns(question: str, all_cols: list[str]) -> tuple[str, list[str]]:
    """Return (group_name, ordered family columns) for MA questions like s10_1_O8."""
    q = str(question).strip()
    m = re.match(r"^(.*)_O(\d+)$", q, flags=re.IGNORECASE)
    if not m:
        return "", []

    prefix = m.group(1)
    family: list[tuple[int, str]] = []
    for col in all_cols:
        m_col = re.match(rf"^{re.escape(prefix)}_O(\d+)$", str(col), flags=re.IGNORECASE)
        if m_col:
            family.append((int(m_col.group(1)), str(col)))

    family.sort(key=lambda x: x[0])
    return prefix, [col for _, col in family]


def _get_ma_family_other_codes(
    question: str,
    all_cols: list[str],
    value_labels: dict[str, dict],
) -> set[str]:
    """Return all detected Other codes across an MA family."""
    _, family_cols = _get_ma_family_columns(question, all_cols)
    other_codes: set[str] = set()
    for col in family_cols:
        for code in detect_other_codes(col, value_labels):
            other_codes.add(_normalize_code_str(code))
    return other_codes


def _build_ma_answer_summary(
    row: pd.Series,
    question: str,
    all_cols: list[str],
    exclude_codes: set[str] | None = None,
) -> str:
    """List all selected answers in the MA family for the current respondent."""
    _, family_cols = _get_ma_family_columns(question, all_cols)
    if not family_cols:
        return ""

    excluded = exclude_codes or set()
    answers: list[str] = []
    seen: set[str] = set()
    for col in family_cols:
        value = row.get(col, "")
        if not _is_meaningful_ma_value(value):
            continue
        normalized = _normalize_code_str(value)
        if normalized in excluded:
            continue
        if normalized in seen:
            continue
        answers.append(normalized)
        seen.add(normalized)
    return ",".join(answers)


def attach_ma_answer_lists(
    coding_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    value_labels: dict[str, dict],
) -> pd.DataFrame:
    """Add a helper column listing all selected MA answers for each coding row."""
    out = coding_df.copy()
    out[MA_CODES_COL] = ""

    if out.empty:
        return out

    sbjnum_source_col = _resolve_sbjnum_col(raw_df)
    if not sbjnum_source_col or SBJNUM_COL not in out.columns:
        return out

    raw_cols = [str(c) for c in raw_df.columns]
    raw_index: dict[str, int] = {
        str(v).strip(): i for i, v in raw_df[sbjnum_source_col].items()
    }
    family_other_codes_cache: dict[str, set[str]] = {}
    for idx, row in out.iterrows():
        question = str(row.get("Question", "")).strip()
        group_name, family_cols = _get_ma_family_columns(question, raw_cols)
        if not group_name or not family_cols:
            continue

        sbjnum = str(row.get(SBJNUM_COL, "")).strip()
        raw_idx = raw_index.get(sbjnum)
        if raw_idx is None:
            continue

        if group_name not in family_other_codes_cache:
            family_other_codes_cache[group_name] = _get_ma_family_other_codes(
                question, raw_cols, value_labels
            )
        exclude_codes = family_other_codes_cache[group_name]
        out.at[idx, MA_CODES_COL] = _build_ma_answer_summary(
            raw_df.loc[raw_idx],
            question,
            raw_cols,
            exclude_codes=exclude_codes,
        )

    return out


def _coerce_new_code_like_old_value(new_code: str, old_value):
    """Preserve numeric cell type when possible for Phase 2 output."""
    s = str(new_code).strip()
    if not s:
        return s

    is_int_text = re.fullmatch(r"[+-]?\d+", s) is not None
    is_float_text = re.fullmatch(r"[+-]?\d+\.\d+", s) is not None

    if pd.isna(old_value) or isinstance(old_value, numbers.Real):
        if is_int_text:
            return int(s)
        if is_float_text:
            return float(s)
    return s


def _clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_empty_value(value) -> bool:
    if pd.isna(value):
        return True
    s = str(value).strip().lower()
    return s in {"", "nan", "none", "null"}


def _text_or_empty(value) -> str:
    return "" if _is_empty_value(value) else str(value).strip()


def _is_cut_command(value: str) -> bool:
    return _text_or_empty(value).lower() == "ตัด"


# ---------------------------------------------------------------------------
# Phase 1 — Build Coding Sheet
# ---------------------------------------------------------------------------

def build_coding_df(
    df: pd.DataFrame,
    pairs: list[tuple[str, str]],
    variable_labels: dict[str, str],
    value_labels: dict[str, dict],
) -> tuple[pd.DataFrame, dict[str, list[str]]]:
    """
    Build the 'Other Coding Sheet' DataFrame.
    Returns (coding_df, detected_codes_map).
    """
    records: list[dict] = []
    detected: dict[str, list[str]] = {}
    sbjnum_source_col = _resolve_sbjnum_col(df)
    cols = [str(c) for c in df.columns]

    # Build wide-group map: prefix -> oth index map and O* columns
    wide_groups: dict[str, dict] = {}
    for c in cols:
        prefix, idx = _parse_oth_col_name(c)
        if prefix is not None and idx is not None:
            g = wide_groups.setdefault(prefix, {"oth": {}, "q": []})
            g["oth"][idx] = c
            continue
        q_prefix, q_idx = _parse_question_family(c)
        if q_prefix is not None and q_idx is not None:
            g = wide_groups.setdefault(q_prefix, {"oth": {}, "q": []})
            g["q"].append(c)

    pair_set = set(pairs)
    q_with_oth: set[str] = {q for q, _ in pairs}
    existing_keys: set[tuple[str, str, str, str, str]] = set()
    existing_open_text_keys: set[tuple[str, str, str]] = set()

    def _add_record(rec: dict) -> None:
        key = (
            str(rec.get("Question", "")),
            str(rec.get(SBJNUM_COL, "")),
            str(rec.get("Other_Code", "")),
            str(rec.get("Open_Text_From", "")),
            str(rec.get("Open_Text", "")),
        )
        if key in existing_keys:
            return
        existing_keys.add(key)
        open_key = (
            str(rec.get(SBJNUM_COL, "")),
            str(rec.get("Open_Text_From", "")),
            str(rec.get("Open_Text", "")),
        )
        existing_open_text_keys.add(open_key)
        records.append(rec)

    def _append_from_mask(
        q_col: str,
        oth_col: str | None,
        code_list: list[str],
        fallback_oth_cols: list[str] | None = None,
        oth_only_code: str | None = None,
    ) -> None:
        if q_col not in df.columns:
            return
        vl = _get_value_labels_for_qcol(q_col, value_labels)
        var_label = variable_labels.get(q_col, "") or ""
        code_norm_set = {_normalize_code_str(c) for c in code_list}
        q_series_norm = df[q_col].astype(str).str.strip().map(_normalize_code_str)
        mask = q_series_norm.isin(code_norm_set)

        extra_mask = pd.Series(False, index=df.index)
        oth_only_code_norm = _normalize_code_str(oth_only_code) if oth_only_code else ""
        if oth_only_code_norm and oth_col and oth_col in df.columns:
            oth_nonempty = df[oth_col].map(lambda v: _clean_text(v) != "")
            q_empty = df[q_col].map(_is_empty_value)
            extra_mask = oth_nonempty & q_empty & (~mask)

        mask_all = mask | extra_mask
        if not mask_all.any():
            return
        subset = df.loc[mask_all]
        logger.info(f"  {q_col}: codes={code_list}, {len(subset)} row(s)")
        for idx, row in subset.iterrows():
            if bool(mask.loc[idx]):
                current_code = _normalize_code_str(df.at[idx, q_col])
            else:
                current_code = oth_only_code_norm
            main_text = _clean_text(row.get(oth_col, "")) if oth_col else ""

            # If mapped oth is empty, only fallback when there is exactly one
            # unambiguous non-empty oth column in the row. Never list all oth texts.
            fallback_items: list[tuple[str, str]] = []
            if not main_text and fallback_oth_cols:
                for c in fallback_oth_cols:
                    txt = _clean_text(row.get(c, ""))
                    if txt:
                        fallback_items.append((c, txt))

            if not main_text and fallback_items:
                if oth_col:
                    fallback_items = [(src_col, txt) for src_col, txt in fallback_items if src_col != oth_col]
                if len(fallback_items) == 1:
                    src_col, txt = fallback_items[0]
                    _add_record(
                        {
                            "Question": q_col,
                            "Variable_Label": var_label,
                            SBJNUM_COL: row.get(sbjnum_source_col, "") if sbjnum_source_col else "",
                            "Other_Code": current_code,
                            "Other_Label": _get_label(vl, current_code),
                            "Open_Text_From": src_col,
                            "Open_Text": txt,
                            NEW_CODE_COL: "",
                            "Remark": "",
                        }
                    )
                    continue
                if len(fallback_items) > 1:
                    logger.info(
                        f"  {q_col}: skip ambiguous oth fallback for sbjnum="
                        f"{row.get(sbjnum_source_col, '') if sbjnum_source_col else ''}"
                    )
                    continue

            if oth_col and not main_text:
                continue

            _add_record(
                {
                    "Question": q_col,
                    "Variable_Label": var_label,
                    SBJNUM_COL: row.get(sbjnum_source_col, "") if sbjnum_source_col else "",
                    "Other_Code": current_code,
                    "Other_Label": _get_label(vl, current_code),
                    "Open_Text_From": oth_col or "",
                    "Open_Text": main_text,
                    NEW_CODE_COL: "",
                    "Remark": "",
                }
            )

    # 1) Process direct pairs
    for q_col, oth_col in pairs:
        other_codes = detect_other_codes(q_col, value_labels)
        if not other_codes:
            logger.info(f"  {q_col}: no other label in SPSS - skip")
            continue

        m_oth_idx = re.match(r"^.*_oth(\d+)$", oth_col, flags=re.IGNORECASE)
        forced_code: str | None = None
        if m_oth_idx:
            forced_code = f"98{int(m_oth_idx.group(1)):02d}"
            if forced_code in other_codes:
                other_codes = [forced_code]

        detected[q_col] = other_codes
        _append_from_mask(
            q_col,
            oth_col,
            other_codes,
            oth_only_code=forced_code if forced_code in other_codes else None,
        )

    # 2) Process wide groups not covered by direct pairs
    for prefix, grp in wide_groups.items():
        oth_map: dict[int, str] = grp["oth"]
        q_cols: list[str] = grp["q"]
        if len(oth_map) == 0 or len(q_cols) < 5:
            continue
        q_with_oth.update(q_cols)

        for q_col in q_cols:
            if any((q_col, oth_col) in pair_set for oth_col in oth_map.values()):
                continue

            other_codes = detect_other_codes(q_col, value_labels)
            if not other_codes:
                continue
            detected[q_col] = other_codes

            mapped_codes: dict[str, str | None] = {}
            for code in other_codes:
                norm_code = _normalize_code_str(code)
                m98 = re.match(r"^98(\d+)$", norm_code)
                if m98:
                    idx = int(m98.group(1))
                    mapped_codes[norm_code] = oth_map.get(idx)
                elif norm_code.isdigit():
                    mapped_codes[norm_code] = oth_map.get(int(norm_code))
                else:
                    mapped_codes[norm_code] = None

            for code_norm, oth_col in mapped_codes.items():
                _append_from_mask(q_col, oth_col, [code_norm], fallback_oth_cols=list(oth_map.values()))

    # 3) Process questions without _oth columns:
    # if SPSS label has "other" code and rawdata contains that code, export it too.
    for q_col in cols:
        if q_col in q_with_oth:
            continue
        if q_col == (sbjnum_source_col or ""):
            continue
        other_codes = detect_other_codes(q_col, value_labels)
        if not other_codes:
            continue
        detected[q_col] = other_codes
        inferred_oth_col = _infer_oth_col_for_question(q_col, cols)
        _append_from_mask(q_col, inferred_oth_col, other_codes)

    # 4) Hard guarantee: pull every row with non-empty *_oth text (any pattern).
    other_codes_cache: dict[str, list[str]] = {}
    for oth_col in cols:
        if not _has_oth_marker(oth_col):
            continue
        if oth_col not in df.columns:
            continue

        oth_text = df[oth_col].map(_clean_text)
        non_empty_mask = oth_text != ""
        if not non_empty_mask.any():
            continue

        q_candidates = _infer_q_candidates_for_oth(oth_col, cols)
        if not q_candidates:
            continue

        oth_prefix, oth_idx = _parse_oth_col_name(oth_col)
        forced_code = None
        if oth_idx is not None:
            if re.search(r"_oth\d+$", oth_col, flags=re.IGNORECASE):
                forced_code = f"98{int(oth_idx):02d}"
            else:
                forced_code = str(int(oth_idx))

        for idx in df.index[non_empty_mask]:
            sbj = (
                str(df.at[idx, sbjnum_source_col]).strip()
                if sbjnum_source_col and sbjnum_source_col in df.columns
                else ""
            )
            txt = oth_text.at[idx]

            chosen_q: str | None = None
            chosen_code: str = ""
            auto_cut_open_text = ""

            # Prefer candidate where row value already equals one of detected other codes.
            for q_col in q_candidates:
                if q_col not in df.columns:
                    continue
                if q_col not in other_codes_cache:
                    other_codes_cache[q_col] = detect_other_codes(q_col, value_labels)
                q_other_codes = {_normalize_code_str(c) for c in other_codes_cache[q_col]}
                if not q_other_codes:
                    continue
                qv = _normalize_code_str(df.at[idx, q_col])
                if qv in q_other_codes:
                    chosen_q = q_col
                    chosen_code = qv
                    break

            # Fallback: if we found a verbatim but no actual Other code was selected,
            # still list it in the coding sheet with blank Other_Code and prefill "ตัด".
            if not chosen_q:
                for q_col in q_candidates:
                    if q_col not in df.columns:
                        continue
                    if q_col not in other_codes_cache:
                        other_codes_cache[q_col] = detect_other_codes(q_col, value_labels)
                    q_other_codes = [_normalize_code_str(c) for c in other_codes_cache[q_col]]
                    if not q_other_codes:
                        continue
                    chosen_q = q_col
                    chosen_code = ""
                    auto_cut_open_text = "ตัด"
                    break

            if not chosen_q:
                continue

            vl = _get_value_labels_for_qcol(chosen_q, value_labels)
            open_key = (sbj, oth_col, txt)
            if open_key in existing_open_text_keys:
                continue
            _add_record(
                {
                    "Question": chosen_q,
                    "Variable_Label": variable_labels.get(chosen_q, "") or "",
                    SBJNUM_COL: sbj,
                    "Other_Code": chosen_code,
                    "Other_Label": _get_label(vl, chosen_code) if chosen_code else "",
                    "Open_Text_From": oth_col,
                    "Open_Text": txt,
                    NEW_CODE_COL: "",
                    NEW_OPEN_TEXT_COL: auto_cut_open_text,
                    "Remark": "",
                }
            )

    return pd.DataFrame(records), detected


def _normalize_coding_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure Coding Sheet has the expected columns in fixed order."""
    out = df.copy()
    for col in CODING_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    out = out[CODING_COLUMNS].copy()
    for col in CODING_COLUMNS:
        out[col] = out[col].map(_text_or_empty)
    return out


def _read_existing_coding_sheet(path: Path) -> pd.DataFrame:
    """Read existing coding workbook (supports new layout with Back-to-Index row)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    frames: list[pd.DataFrame] = []
    for ws in wb.worksheets:
        if ws.title == "Index":
            continue
        first_cell = ws.cell(row=1, column=1).value
        header_row = 1
        if str(first_cell or "").strip().lower() == "back to index":
            header_row = 2
        try:
            part = pd.read_excel(path, sheet_name=ws.title, dtype=str, header=header_row - 1)
        except Exception:
            continue
        if part is None or part.empty:
            continue
        part = _normalize_coding_columns(part)
        frames.append(part)
    if not frames:
        return pd.DataFrame(columns=CODING_COLUMNS)
    return pd.concat(frames, ignore_index=True)


def _merge_coding_with_existing(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """
    Append new items to existing coding sheet by key, preserving old edits.
    Existing rows are kept first so existing New_Code/Remark remain unchanged.
    """
    existing_norm = _normalize_coding_columns(existing_df)
    new_norm = _normalize_coding_columns(new_df)
    existing_norm["_merge_order"] = range(len(existing_norm))
    new_norm["_merge_order"] = range(len(existing_norm), len(existing_norm) + len(new_norm))
    key_cols = ["Question", SBJNUM_COL, "Other_Code", "Open_Text_From", "Open_Text"]
    merged = pd.concat([existing_norm, new_norm], ignore_index=True)
    merged = merged.drop_duplicates(subset=key_cols, keep="first")
    return merged


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Return Excel-safe unique sheet name (max 31 chars)."""
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name)).strip() or "Question"
    base = base[:31]
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = f"{base[: max(1, 31 - len(suffix))]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


def _sheet_group_for_question(question: str) -> str:
    """Group MA columns like s10_1_O1..O31 into one sheet named by prefix (s10_1)."""
    q = str(question).strip()
    m = re.match(r"^(.*)_O\d+$", q, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    return q


def _ordered_sheet_groups(coding_df: pd.DataFrame, source_columns: list[str] | None) -> list[str]:
    """Order sheet groups by original rawdata column order when available."""
    groups = coding_df["Question"].map(_sheet_group_for_question).dropna().astype(str).tolist()
    if not groups:
        return []
    present = set(groups)
    ordered: list[str] = []
    seen: set[str] = set()

    if source_columns:
        for col in source_columns:
            g = _sheet_group_for_question(str(col))
            if g in present and g not in seen:
                ordered.append(g)
                seen.add(g)

    for g in groups:
        if g not in seen:
            ordered.append(g)
            seen.add(g)
    return ordered


def _sort_coding_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Sort coding rows to keep same Sbjnum grouped together."""
    out = df.copy()
    if "_merge_order" in out.columns:
        out = out.sort_values(by=["_merge_order"], kind="mergesort")
        return out.drop(columns=["_merge_order"])
    if SBJNUM_COL in out.columns:
        sbj_series = out[SBJNUM_COL].astype(str).str.strip()
        sbj_num = pd.to_numeric(sbj_series, errors="coerce")
        out["_sbj_is_num"] = sbj_num.notna().map(lambda x: 0 if x else 1)
        out["_sbj_num"] = sbj_num.fillna(0)
        out["_sbj_str"] = sbj_series
    else:
        out["_sbj_is_num"] = 1
        out["_sbj_num"] = 0
        out["_sbj_str"] = ""

    oc_series = out["Other_Code"].astype(str).str.strip()
    oc_num = pd.to_numeric(oc_series, errors="coerce")
    out["_oc_is_num"] = oc_num.notna().map(lambda x: 0 if x else 1)
    out["_oc_num"] = oc_num.fillna(0)
    out["_oc_str"] = oc_series
    out["_q"] = out["Question"].astype(str)
    out = out.sort_values(
        by=[
            "_sbj_is_num", "_sbj_num", "_sbj_str",
            "_oc_is_num", "_oc_num", "_oc_str",
            "_q",
        ],
        kind="mergesort",
    )
    return out.drop(
        columns=[
            "_sbj_is_num", "_sbj_num", "_sbj_str",
            "_oc_is_num", "_oc_num", "_oc_str",
            "_q",
        ]
    )


def save_coding_sheet(
    coding_df: pd.DataFrame,
    output_path: Path,
    source_columns: list[str] | None = None,
) -> None:
    """Save coding_df to Excel (group MA _O* into one sheet) with styling."""
    used_sheet_names: set[str] = {"Index"}
    tmp = coding_df.copy()
    tmp["_sheet_group"] = tmp["Question"].map(_sheet_group_for_question)

    sheet_groups = _ordered_sheet_groups(tmp, source_columns)
    sheet_items: list[tuple[str, str, pd.DataFrame]] = []
    for group_name in sheet_groups:
        sub_df = tmp[tmp["_sheet_group"] == group_name].drop(columns=["_sheet_group"])
        sub_df = _sort_coding_rows(sub_df)
        ma_header = f"Code {str(group_name).upper()}(MA)"
        has_ma_values = (
            MA_CODES_COL in sub_df.columns
            and sub_df[MA_CODES_COL].fillna("").astype(str).str.strip().ne("").any()
        )
        if has_ma_values:
            cols = [c for c in sub_df.columns if c != MA_CODES_COL]
            if "Other_Code" in cols:
                insert_at = cols.index("Other_Code")
                cols.insert(insert_at, MA_CODES_COL)
            else:
                cols.append(MA_CODES_COL)
            sub_df = sub_df[cols].rename(columns={MA_CODES_COL: ma_header})
        elif MA_CODES_COL in sub_df.columns:
            sub_df = sub_df.drop(columns=[MA_CODES_COL])
        sheet_name = _safe_sheet_name(str(group_name), used_sheet_names)
        sheet_items.append((str(group_name), sheet_name, sub_df))

    index_df = pd.DataFrame(
        {
            "No": list(range(1, len(sheet_items) + 1)),
            "QNR": [sheet_name for _, sheet_name, _ in sheet_items],
            "ข้อOther": [
                (
                    "Y"
                    if (
                        "Open_Text" in sub_df.columns
                        and sub_df["Open_Text"].fillna("").astype(str).str.strip().ne("").any()
                    )
                    else "-"
                )
                for _, _, sub_df in sheet_items
            ],
            "Count": [len(sub_df) for _, _, sub_df in sheet_items],
        }
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        index_df.to_excel(writer, sheet_name="Index", index=False)
        for _, sheet_name, sub_df in sheet_items:
            sub_df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_coding_sheet(output_path)
    logger.info(f"Coding sheet saved: {output_path.name}")


def _style_coding_sheet(output_path: Path) -> None:
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        if ws.title == "Index":
            hdr_fill = PatternFill("solid", fgColor="2E4057")
            hdr_font = Font(color="FFFFFF", bold=True, size=11)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            max_row = ws.max_row
            if max_row >= 2:
                table_ref = f"A1:D{max_row}"
                table = Table(displayName="IndexTable", ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(table)

                for r in range(2, max_row + 1):
                    qnr_cell = ws.cell(row=r, column=2)
                    target_sheet = str(qnr_cell.value or "").strip()
                    if target_sheet and target_sheet in wb.sheetnames:
                        qnr_cell.hyperlink = f"#'{target_sheet}'!A1"
                        qnr_cell.font = Font(color="0563C1", underline="single")

                    other_cell = ws.cell(row=r, column=3)
                    other_val = str(other_cell.value or "").strip().upper()
                    if other_val == "Y":
                        other_cell.fill = PatternFill("solid", fgColor="92D050")
                        other_cell.alignment = Alignment(horizontal="center")
                    elif other_val == "-":
                        other_cell.fill = PatternFill("solid", fgColor="FF0000")
                        other_cell.font = Font(color="FFFFFF", bold=True)
                        other_cell.alignment = Alignment(horizontal="center")

            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 8
            ws.row_dimensions[1].height = 24
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
                ws.cell(row=r, column=2).font = Font(color="0563C1", underline="single", bold=True)
                ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

            ws["F1"] = "'Y' = มี Verbatim"
            ws["F2"] = "'-' = ไม่มี Verbatim"
            ws["F1"].font = Font(color="FF0000", bold=True)
            ws["F2"].font = Font(color="FF0000", bold=True)
            continue

        # Reserve top row for navigation link.
        ws.insert_rows(1)

        col_idx: dict[str, int] = {}
        for idx, cell in enumerate(ws[2], start=1):
            if cell.value:
                col_idx[str(cell.value)] = idx

        new_code_idx = col_idx.get(NEW_CODE_COL)
        new_open_text_idx = col_idx.get(NEW_OPEN_TEXT_COL)
        open_text_idx = col_idx.get("Open_Text")

        hdr_fill = PatternFill("solid", fgColor="2E4057")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        yellow = PatternFill("solid", fgColor="FFF176")
        for edit_idx in [new_code_idx, new_open_text_idx]:
            if edit_idx:
                for row in ws.iter_rows(min_row=3, min_col=edit_idx, max_col=edit_idx):
                    for cell in row:
                        cell.fill = yellow

        if new_code_idx:
            ma_idx = None
            for col_name, idx in col_idx.items():
                if re.fullmatch(r"Code .+\(MA\)", col_name):
                    ma_idx = idx
                    break

            other_code_idx = col_idx.get("Other_Code")
            if ws.max_row >= 3 and (ma_idx or other_code_idx):
                new_code_col = get_column_letter(new_code_idx)
                rules: list[str] = []
                if other_code_idx:
                    other_code_col = get_column_letter(other_code_idx)
                    rules.append(
                        f'TRIM(${new_code_col}3)=TRIM(${other_code_col}3)'
                    )
                if ma_idx:
                    ma_col = get_column_letter(ma_idx)
                    rules.append(
                        f'ISNUMBER(SEARCH(","&TRIM(${new_code_col}3)&",",","&SUBSTITUTE(${ma_col}3," ","")&","))'
                    )
                if rules:
                    red_fill = PatternFill("solid", fgColor="FF1F1F")
                    red_font = Font(color="FFFFFF", bold=True)
                    formula = (
                        f'=AND(TRIM(${new_code_col}3)<>"",TRIM(${new_code_col}3)<>"ตัด",'
                        f'OR({",".join(rules)}))'
                    )
                    ws.conditional_formatting.add(
                        f"{new_code_col}3:{new_code_col}{ws.max_row}",
                        FormulaRule(formula=[formula], fill=red_fill, font=red_font),
                    )

                    valid_formula = (
                        f'=OR(TRIM(${new_code_col}3)="",TRIM(${new_code_col}3)="ตัด",'
                        f'NOT(OR({",".join(rules)})))'
                    )
                    dv = DataValidation(
                        type="custom",
                        formula1=valid_formula,
                        allow_blank=True,
                    )
                    dv.showErrorMessage = True
                    dv.errorStyle = "stop"
                    dv.errorTitle = "Duplicate Code"
                    dv.error = (
                        "New_Code ซ้ำกับ code เดิมในแถวนี้ "
                        "(ซ้ำกับ Other_Code หรืออยู่ใน Code(MA))"
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{new_code_col}3:{new_code_col}{ws.max_row}")

        fills = [PatternFill("solid", fgColor="E3F2FD"), PatternFill("solid", fgColor="FFFFFF")]
        for r_idx, row in enumerate(ws.iter_rows(min_row=3), start=0):
            fill = fills[r_idx % 2]
            for cell in row:
                if cell.column not in {new_code_idx, new_open_text_idx}:
                    cell.fill = fill
            if open_text_idx:
                row[open_text_idx - 1].alignment = Alignment(wrap_text=True)

        # quick navigation back to summary sheet
        ws["A1"] = "Back to Index"
        ws["A1"].hyperlink = "#'Index'!A1"
        ws["A1"].font = Font(color="0563C1", underline="single", bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A3"
        width_map = {
            "Question": 14, "Variable_Label": 35, SBJNUM_COL: 12,
            "Other_Code": 12, "Other_Label": 18, "Open_Text_From": 18, "Open_Text": 30,
            NEW_CODE_COL: 14, NEW_OPEN_TEXT_COL: 30, "Remark": 25,
        }
        for col_name, width in width_map.items():
            if col_name in col_idx:
                ws.column_dimensions[get_column_letter(col_idx[col_name])].width = width
        for col_name, idx in col_idx.items():
            if re.fullmatch(r"Code .+\(MA\)", col_name):
                ws.column_dimensions[get_column_letter(idx)].width = 18
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 30
    wb.save(output_path)


def phase1_export(
    rawdata_path: Path,
    spss_path: Path,
    output_path: Path,
    allow_order_fallback: bool = False,
) -> Phase1Result:
    """
    Full Phase 1 pipeline.
    Returns Phase1Result (raises ValueError if nothing to export).
    """
    df = read_rawdata(rawdata_path, as_str=True)
    pairs = find_oth_pairs(df)
    if not pairs:
        raise ValueError("ไม่พบคู่คอลัมน์ _oth ใน Rawdata")

    var_labels, val_labels = read_spss_labels(spss_path)
    mapping_report = _build_spss_column_mapping(
        [str(col) for col in df.columns if str(col) != (_resolve_sbjnum_col(df) or "") and not _has_oth_marker(str(col))],
        var_labels,
        val_labels,
        allow_order_fallback=allow_order_fallback,
    )
    coding_df, detected = build_coding_df(
        df,
        pairs,
        mapping_report.resolved_var_labels,
        mapping_report.resolved_value_labels,
    )

    if coding_df.empty:
        raise ValueError("ไม่พบแถวที่ตอบ Other ในข้อใดเลย")

    if output_path.exists():
        try:
            existing_df = _read_existing_coding_sheet(output_path)
            coding_df = _merge_coding_with_existing(existing_df, coding_df)
            logger.info(f"  -> merge with existing coding sheet: {output_path.name}")
        except Exception as ex:
            logger.warning(
                f"Cannot read existing coding sheet ({output_path.name}), overwrite instead: {ex}"
            )
            coding_df = _normalize_coding_columns(coding_df)
    else:
        coding_df = _normalize_coding_columns(coding_df)

    coding_df = attach_ma_answer_lists(coding_df, df, val_labels)
    save_coding_sheet(coding_df, output_path, source_columns=[str(c) for c in df.columns])
    coding_df = coding_df.drop(columns=["_merge_order"], errors="ignore")
    n_sheet_questions = coding_df["Question"].map(_sheet_group_for_question).nunique()

    return Phase1Result(
        coding_df=coding_df,
        output_path=output_path,
        n_questions=n_sheet_questions,
        n_rows=len(coding_df),
        detected=detected,
    )


# ---------------------------------------------------------------------------
# Phase 2 — Apply Recodes
# ---------------------------------------------------------------------------

def phase2_apply(
    rawdata_path: Path,
    coding_sheet_path: Path,
    output_rawdata_path: Path,
) -> Phase2Result:
    """
    Full Phase 2 pipeline.
    Returns Phase2Result.
    """
    df = read_rawdata(rawdata_path, as_str=False)
    coding_df = _read_existing_coding_sheet(coding_sheet_path)

    required = {"Question", SBJNUM_COL, NEW_CODE_COL}
    missing = required - set(coding_df.columns)
    if missing:
        raise ValueError(f"Coding sheet ขาด column: {missing}")

    if NEW_OPEN_TEXT_COL not in coding_df.columns:
        coding_df[NEW_OPEN_TEXT_COL] = ""

    has_edit = (
        ~coding_df[NEW_CODE_COL].map(_is_empty_value)
        | ~coding_df[NEW_OPEN_TEXT_COL].map(_is_empty_value)
    )
    coded_df = coding_df[has_edit].copy()
    skipped_df = coding_df[~has_edit].copy()

    sbjnum_source_col = _resolve_sbjnum_col(df)
    if not sbjnum_source_col:
        raise ValueError(f"ไม่พบ column '{SBJNUM_COL}' (หรือชื่อใกล้เคียง) ใน Rawdata")

    sbjnum_index: dict[str, int] = {
        str(v).strip(): i for i, v in df[sbjnum_source_col].items()
    }

    log_records: list[dict] = []
    not_found_records: list[dict] = []
    stamped_cells: list[tuple[int, str]] = []  # (df row index, column name)

    for _, crow in coded_df.iterrows():
        q_col = _text_or_empty(crow.get("Question", ""))
        sbjnum = _text_or_empty(crow.get(SBJNUM_COL, ""))
        new_code = _text_or_empty(crow.get(NEW_CODE_COL, ""))
        new_open_text = _text_or_empty(crow.get(NEW_OPEN_TEXT_COL, ""))
        open_text = _text_or_empty(crow.get("Open_Text", ""))
        open_text_from = _text_or_empty(crow.get("Open_Text_From", ""))
        other_code = _text_or_empty(crow.get("Other_Code", ""))
        other_label = _text_or_empty(crow.get("Other_Label", ""))

        row_idx = sbjnum_index.get(sbjnum)
        if row_idx is None:
            not_found_records.append(
                {SBJNUM_COL: sbjnum, "Question": q_col, "Reason": "sbjnum not found"}
            )
            continue

        applied_code = False
        applied_open_text = False
        old_val = ""
        old_open_text = ""
        open_text_target = ""

        cut_code = _is_cut_command(new_code)
        cut_open_text = _is_cut_command(new_open_text)

        if new_code:
            if q_col not in df.columns:
                not_found_records.append(
                    {SBJNUM_COL: sbjnum, "Question": q_col, "Reason": "column not found"}
                )
            else:
                old_cell_value = df.at[row_idx, q_col]
                old_val = str(old_cell_value).strip()
                if cut_code:
                    if not _is_empty_value(old_cell_value):
                        df.at[row_idx, q_col] = ""
                        applied_code = True
                        stamped_cells.append((row_idx, q_col))
                else:
                    df.at[row_idx, q_col] = _coerce_new_code_like_old_value(new_code, old_cell_value)
                    applied_code = True
                    stamped_cells.append((row_idx, q_col))

        if new_open_text or open_text:
            if open_text_from and open_text_from in df.columns:
                open_text_target = open_text_from
            else:
                inferred_oth = _infer_oth_col_for_question(q_col, [str(c) for c in df.columns])
                if inferred_oth and inferred_oth in df.columns:
                    open_text_target = inferred_oth

            if open_text_target:
                old_open_text = str(df.at[row_idx, open_text_target]).strip()
                target_open_text = new_open_text
                if cut_open_text:
                    target_open_text = ""
                if not cut_open_text and _is_empty_value(target_open_text):
                    target_open_text = old_open_text
                if target_open_text != old_open_text:
                    df.at[row_idx, open_text_target] = target_open_text
                    applied_open_text = True
                    stamped_cells.append((row_idx, open_text_target))
                if cut_open_text and q_col in df.columns:
                    old_cell_value = df.at[row_idx, q_col]
                    old_val = str(old_cell_value).strip()
                    if _normalize_code_str(old_cell_value) == _normalize_code_str(other_code):
                        df.at[row_idx, q_col] = ""
                        if (row_idx, q_col) not in stamped_cells:
                            stamped_cells.append((row_idx, q_col))
                        applied_code = True
            elif not applied_code:
                not_found_records.append(
                    {
                        SBJNUM_COL: sbjnum,
                        "Question": q_col,
                        "Reason": "open_text target column not found",
                    }
                )
                continue

        if not applied_code and not applied_open_text:
            continue

        log_records.append(
            {
                SBJNUM_COL: sbjnum,
                "Question": q_col,
                "Old_Code": old_val,
                "Old_Label": other_label,
                "New_Code": new_code,
                NEW_OPEN_TEXT_COL: new_open_text,
                "Open_Text_Column": open_text_target,
                "Old_Open_Text": old_open_text,
                "Open_Text": open_text,
                "Applied_Code": "Y" if applied_code else "",
                "Applied_Open_Text": "Y" if applied_open_text else "",
                "Recoded_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    # Save rawdata
    df.to_excel(output_rawdata_path, index=False)
    _highlight_stamped_cells(output_rawdata_path, df, stamped_cells)
    logger.info(f"✓ Rawdata → {output_rawdata_path.name}")

    log_df = pd.DataFrame(log_records)
    not_found_df = pd.DataFrame(not_found_records)

    return Phase2Result(
        log_df=log_df,
        skipped_df=skipped_df,
        not_found_df=not_found_df,
        output_rawdata_path=output_rawdata_path,
        n_applied=len(log_records),
        n_skipped=len(skipped_df),
        n_not_found=len(not_found_records),
    )


def _highlight_stamped_cells(
    output_rawdata_path: Path,
    df: pd.DataFrame,
    stamped_cells: list[tuple[int, str]],
) -> None:
    """Highlight stamped cells in output rawdata workbook."""
    if not stamped_cells:
        return
    wb = load_workbook(output_rawdata_path)
    ws = wb.active
    col_pos = {str(col): i + 1 for i, col in enumerate(df.columns)}
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    for df_row_idx, col_name in stamped_cells:
        col_idx = col_pos.get(str(col_name))
        if not col_idx:
            continue
        excel_row = int(df_row_idx) + 2  # header is row 1
        ws.cell(row=excel_row, column=col_idx).fill = green_fill
    wb.save(output_rawdata_path)


def _style_log(log_path: Path) -> None:
    wb = load_workbook(log_path)
    colors = {"Recode_Log": "2E7D32", "Not_Found": "B71C1C", "Skipped_No_Code": "E65100"}
    for ws in wb.worksheets:
        color = colors.get(ws.title, "388E3C")
        hdr_fill = PatternFill("solid", fgColor=color)
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)
    wb.save(log_path)


def _extract_json_object(text: str) -> dict:
    """Extract the first JSON object from a model response."""
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("AI returned empty content")

    fenced = re.search(r"```(?:json)?\s*(\{.*\})\s*```", raw, flags=re.DOTALL)
    if fenced:
        raw = fenced.group(1).strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}")
        if start >= 0 and end > start:
            return json.loads(raw[start : end + 1])
        raise


def _build_codeframe_prompt(
    group_name: str,
    variable_label: str,
    items: list[dict[str, object]],
) -> str:
    """Prompt the AI to cluster open text into a codeframe."""
    lines = [
        f"{int(item['id'])}. {item['text']} (count={int(item['count'])})"
        for item in items
    ]
    joined_texts = "\n".join(lines)
    return (
        "You are a market research coding expert.\n"
        "Create a concise summary codeframe from the respondent open-text answers.\n"
        "Group similar meanings together, normalize spelling variations, and keep categories mutually exclusive.\n"
        "Return JSON only with this exact shape:\n"
        "{\n"
        '  "categories": [\n'
        "    {\n"
        '      "code": "1",\n'
        '      "thai_group2": "Thai grouped label",\n'
        '      "english": "short English label",\n'
        '      "matched_item_ids": [1, 2, 3]\n'
        "    }\n"
        "  ]\n"
        "}\n"
        "Rules:\n"
        "- Use short numeric codes starting from 1.\n"
        "- thai_group2 should be the grouped Thai meaning for similar answers.\n"
        "- english should be short and simple.\n"
        "- Put each input item id in the single best category.\n"
        "- Every input item id should appear once across all categories.\n"
        "- Do not invent item ids.\n"
        "- Keep 4 to 12 categories unless the data clearly needs fewer.\n"
        "- Keep the response very compact.\n"
        f"Question group: {group_name}\n"
        f"Variable label: {variable_label}\n"
        "Open-text answers:\n"
        f"{joined_texts}"
    )


def _call_openrouter_chat(api_key: str, model: str, prompt: str) -> dict:
    """Call OpenRouter chat completions and parse JSON response."""
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": "You produce strict JSON only.",
            },
            {
                "role": "user",
                "content": prompt,
            },
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
    }
    req = urlrequest.Request(
        OPENROUTER_BASE_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://local.other-tool",
            "X-Title": "Tools Other CE V1",
        },
        method="POST",
    )
    try:
        with urlrequest.urlopen(req, timeout=120) as resp:
            body = resp.read().decode("utf-8")
    except urlerror.HTTPError as ex:
        detail = ex.read().decode("utf-8", errors="replace")
        message = f"OpenRouter HTTP {ex.code}"
        try:
            error_body = json.loads(detail)
            error_info = error_body.get("error", {})
            error_message = str(error_info.get("message", "")).strip()
            if error_message:
                message = f"{message}: {error_message}"
        except Exception:
            trimmed = detail.strip()
            if trimmed:
                message = f"{message}: {trimmed[:300]}"
        if ex.code == 404:
            message = (
                f"{message}\n\n"
                "ตรวจสอบชื่อ model ในหน้า AI Phase 3 อีกครั้ง เช่น `openrouter/auto` "
                "หรือ model slug ที่มีอยู่จริงบน OpenRouter"
            )
        raise ValueError(message) from ex
    except urlerror.URLError as ex:
        raise ValueError(f"OpenRouter connection failed: {ex}") from ex

    data = json.loads(body)
    try:
        content = data["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as ex:
        raise ValueError(f"Unexpected OpenRouter response: {body[:500]}") from ex
    return _extract_json_object(content)


def _normalize_codeframe_rows(
    group_name: str,
    variable_label: str,
    categories: list[dict],
    items: list[dict[str, object]],
    model: str,
) -> list[dict]:
    """Normalize AI categories into workbook rows."""
    item_map = {
        int(item["id"]): {
            "text": str(item["text"]).strip(),
            "count": int(item["count"]),
        }
        for item in items
    }
    rows: list[dict] = []
    for idx, cat in enumerate(categories, start=1):
        matched_ids = [
            int(x) for x in cat.get("matched_item_ids", [])
            if str(x).strip().isdigit() and int(x) in item_map
        ]
        total_count = sum(int(item_map[i]["count"]) for i in matched_ids)
        matched_texts = [str(item_map[i]["text"]).strip() for i in matched_ids if str(item_map[i]["text"]).strip()]
        thai_group1 = " / ".join(matched_texts)
        thai_group2 = str(cat.get("thai_group2", "")).strip()
        if not thai_group1 and thai_group2:
            thai_group1 = thai_group2
        if not thai_group2 and thai_group1:
            thai_group2 = thai_group1
        rows.append(
            {
                "Source_Group": group_name,
                "Variable_Label": variable_label,
                "Code No.": str(cat.get("code", idx)).strip() or str(idx),
                "Thai Group1": thai_group1 or f"Category {idx}",
                "Thai Group2": thai_group2 or thai_group1 or f"Category {idx}",
                "English": str(cat.get("english", "")).strip(),
                "Count": total_count,
                "AI_Model": model,
            }
        )
    return rows


def _save_codeframe_workbook(
    codeframe_df: pd.DataFrame,
    output_path: Path,
    sheet_groups: list[str] | None = None,
) -> None:
    """Save AI-generated codeframe workbook using the requested template layout."""
    tmp = codeframe_df.copy()
    tmp["_sheet_group"] = tmp["Source_Group"].astype(str)
    if sheet_groups is None:
        sheet_groups = _ordered_sheet_groups(
            tmp.rename(columns={"Source_Group": "Question"}),
            source_columns=tmp["Source_Group"].astype(str).drop_duplicates().tolist(),
        )

    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    index_ws = default_ws
    index_ws.title = "Index"
    used_sheet_names: set[str] = {"Index"}

    green_fill = PatternFill("solid", fgColor="C6E0B4")
    hdr_fill = PatternFill("solid", fgColor="DCE6F1")
    index_fill = PatternFill("solid", fgColor="2E4057")
    thin_side = Side(style="thin", color="7F7F7F")

    index_rows: list[tuple[int, str, str, int]] = []

    for seq, group_name in enumerate(sheet_groups, start=1):
        sub_df = (
            tmp[tmp["_sheet_group"] == group_name]
            .drop(columns=["_sheet_group"])
            .reset_index(drop=True)
        )
        sheet_name = _safe_sheet_name(group_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_format.defaultRowHeight = 15
        variable_label = str(sub_df["Variable_Label"].iloc[0] if not sub_df.empty else "").strip()
        index_rows.append((seq, sheet_name, variable_label, len(sub_df)))

        ws["A1"] = "Back to Index"
        ws["A1"].hyperlink = "#'Index'!A1"
        ws["A1"].font = Font(color="0563C1", underline="single", bold=True, size=11)

        ws["A2"] = str(group_name)
        ws["B2"] = variable_label
        ws.merge_cells("B2:D2")

        ws["A2"].fill = green_fill
        ws["B2"].fill = green_fill
        ws["A2"].font = Font(bold=True, size=10)
        ws["B2"].font = Font(bold=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].alignment = Alignment(vertical="center", wrap_text=False)
        for cell_ref in ("A2", "B2", "C2", "D2"):
            ws[cell_ref].border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        headers = ["Code No.", "Thai Group1", "Thai Group2", "English", "Count"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = hdr_fill
            cell.font = Font(bold=True, size=10, color="000000" if header != "Count" else "0000FF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        start_row = 5
        for row_idx, (_, row) in enumerate(sub_df.iterrows(), start=start_row):
            ws.cell(row=row_idx, column=1, value=row["Code No."]).font = Font(size=10)
            ws.cell(row=row_idx, column=2, value=row["Thai Group1"]).font = Font(size=10)
            ws.cell(row=row_idx, column=3, value=row["Thai Group2"]).font = Font(size=10)
            ws.cell(row=row_idx, column=4, value=row["English"]).font = Font(size=10)
            ws.cell(row=row_idx, column=5, value=row["Count"]).font = Font(size=10, color="0000FF")

        max_row = max(38, start_row + len(sub_df) - 1)
        for row in ws.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=5):
            for cell in row:
                if cell.row >= 5:
                    if cell.column in (1, 5):
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side,
                )

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 8
        ws.row_dimensions[1].height = 15
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        for fixed_row in range(5, max_row + 1):
            ws.row_dimensions[fixed_row].height = 15
        ws.freeze_panes = "A5"

    index_ws["A1"] = "No"
    index_ws["B1"] = "Codeframe"
    index_ws["C1"] = "Question"
    index_ws["D1"] = "Rows"
    index_ws.sheet_format.defaultRowHeight = 15
    for cell in index_ws[1]:
        cell.fill = index_fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row_idx, (seq, sheet_name, variable_label, row_count) in enumerate(index_rows, start=2):
        index_ws.cell(row=row_idx, column=1, value=seq).font = Font(size=10)
        link_cell = index_ws.cell(row=row_idx, column=2, value=sheet_name)
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="0563C1", underline="single", size=10)
        index_ws.cell(row=row_idx, column=3, value=variable_label).font = Font(size=10)
        index_ws.cell(row=row_idx, column=4, value=row_count).font = Font(size=10)
        for cell in index_ws[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        index_ws.row_dimensions[row_idx].height = 15

    index_ws.freeze_panes = "A2"
    index_ws.column_dimensions["A"].width = 8
    index_ws.column_dimensions["B"].width = 18
    index_ws.column_dimensions["C"].width = 42
    index_ws.column_dimensions["D"].width = 10
    index_ws.row_dimensions[1].height = 15

    wb.save(output_path)
    logger.info(f"Codeframe saved: {output_path.name}")


def _generate_single_group_codeframe(
    token: str,
    model: str,
    group_name: str,
    group_df: pd.DataFrame,
) -> list[dict]:
    """Generate codeframe rows for a single question group."""
    variable_label = (
        group_df["Variable_Label"].fillna("").astype(str).str.strip().iloc[0]
        if "Variable_Label" in group_df.columns
        else ""
    )
    text_counts = (
        group_df["Open_Text"]
        .value_counts(sort=True)
        .rename_axis("Open_Text")
        .reset_index(name="Count")
    )
    items = [
        {
            "id": i + 1,
            "text": str(row["Open_Text"]).strip(),
            "count": int(row["Count"]),
        }
        for i, (_, row) in enumerate(text_counts.iterrows())
        if str(row["Open_Text"]).strip()
    ]
    if not items:
        return []

    logger.info(f"Generating AI codeframe for {group_name} ({len(items)} unique texts)")
    prompt = _build_codeframe_prompt(group_name, variable_label, items)
    ai_json = _call_openrouter_chat(token, model, prompt)
    categories = ai_json.get("categories")
    if not isinstance(categories, list) or not categories:
        raise ValueError(f"AI returned no categories for {group_name}")

    return _normalize_codeframe_rows(group_name, variable_label, categories, items, model)


def generate_codeframe_with_ai(
    coding_sheet_path: Path,
    output_path: Path,
    api_key: str | None = None,
    model: str = OPENROUTER_MODEL_DEFAULT,
) -> CodeFrameResult:
    """Build a CodeFrame workbook from coding sheet open text using OpenRouter."""
    token = (api_key or os.getenv("OPENROUTER_API_KEY", "")).strip()
    if not token:
        raise ValueError("OpenRouter API key is required")

    coding_df = _read_existing_coding_sheet(coding_sheet_path)
    if coding_df.empty:
        raise ValueError("Coding sheet is empty")

    working = coding_df.copy()
    working["Open_Text"] = working["Open_Text"].fillna("").astype(str).str.strip()
    working = working[working["Open_Text"] != ""].copy()
    if working.empty:
        raise ValueError("No Open_Text rows found in coding sheet")

    working["Source_Group"] = working["Question"].map(_sheet_group_for_question)
    grouped_items = [
        (str(group_name), group_df.copy())
        for group_name, group_df in working.groupby("Source_Group", sort=False)
    ]
    group_order = [group_name for group_name, _ in grouped_items]

    max_workers = min(OPENROUTER_MAX_WORKERS, max(1, len(grouped_items)))
    rows_by_group: dict[str, list[dict]] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_generate_single_group_codeframe, token, model, group_name, group_df): group_name
            for group_name, group_df in grouped_items
        }
        for future in as_completed(futures):
            group_name = futures[future]
            try:
                rows = future.result()
            except Exception as ex:
                raise ValueError(f"CodeFrame failed for {group_name}: {ex}") from ex
            rows_by_group[group_name] = rows

    all_rows: list[dict] = []
    for group_name in group_order:
        all_rows.extend(rows_by_group.get(group_name, []))

    if not all_rows:
        raise ValueError("AI codeframe generation returned no rows")

    codeframe_df = pd.DataFrame(all_rows, columns=CODEFRAME_COLUMNS)
    _save_codeframe_workbook(codeframe_df, output_path, sheet_groups=group_order)
    return CodeFrameResult(
        codeframe_df=codeframe_df,
        output_path=output_path,
        n_groups=len({str(row["Source_Group"]) for row in all_rows}),
        n_rows=len(codeframe_df),
        model=model,
    )
